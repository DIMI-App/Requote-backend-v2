import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from collections import OrderedDict
import openai

openai.api_key = os.environ.get('OPENAI_API_KEY')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OFFER_2_XLSX = os.path.join(BASE_DIR, "offer2_template.xlsx")
OFFER_2_XLS = os.path.join(BASE_DIR, "offer2_template.xls")
ITEMS_PATH = os.path.join(BASE_DIR, "outputs", "items_offer1.json")
OUTPUT_PATH = os.path.join(BASE_DIR, "outputs", "final_offer1.xlsx")

print("=" * 60)
print("GENERATE OFFER FROM XLSX - SV14 with Technical Sections")
print("=" * 60)

# Determine which template exists
if os.path.exists(OFFER_2_XLSX):
    OFFER_2_PATH = OFFER_2_XLSX
elif os.path.exists(OFFER_2_XLS):
    OFFER_2_PATH = OFFER_2_XLS
else:
    print("✗ No XLSX template found")
    exit(1)

# Load items AND technical sections
try:
    with open(ITEMS_PATH, "r", encoding="utf-8") as f:
        full_data = json.load(f)
    items = full_data.get("items", [])
    technical_sections = full_data.get("technical_sections", [])
    print(f"✓ Loaded {len(items)} items")
    print(f"✓ Loaded {len(technical_sections)} technical sections")
except Exception as e:
    print(f"✗ Error loading data: {str(e)}")
    exit(1)

# Load Excel template
try:
    workbook = openpyxl.load_workbook(OFFER_2_PATH)
    sheet = workbook.active
    print(f"✓ Template loaded: {sheet.title}")
except Exception as e:
    print(f"✗ Error loading template: {str(e)}")
    exit(1)

# Insert technical sections before pricing table
if technical_sections and len(technical_sections) > 0:
    print(f"=" * 60)
    print(f"INSERTING {len(technical_sections)} TECHNICAL SECTIONS")
    print(f"=" * 60)
    
    # Find the pricing table header row
    table_header_row = None
    for row_idx in range(1, min(20, sheet.max_row + 1)):
        row = sheet[row_idx]
        row_values = [str(cell.value).upper() if cell.value else '' for cell in row]
        row_text = ' '.join(row_values)
        
        if any(kw in row_text for kw in ['POSITION', 'DESCRIPTION', 'PRICE', 'QUANTITY', 'TOTAL']):
            table_header_row = row_idx
            print(f"✓ Found pricing table header at row {table_header_row}")
            break
    
    if table_header_row:
        # Insert technical sections BEFORE the pricing table
        current_row = table_header_row
        sections_inserted = 0
        
        for section in technical_sections:
            location = section.get('page_location', '')
            
            # Only insert sections that should appear before price table
            if location == 'before_price_table' or location == '':
                title = section.get('section_title', '')
                content = section.get('content', '')
                content_type = section.get('content_type', 'paragraph')
                
                if not content:
                    continue
                
                # Insert blank row
                sheet.insert_rows(current_row)
                
                # Add section title
                if title:
                    title_cell = sheet.cell(current_row, 1)
                    title_cell.value = title
                    title_cell.font = Font(bold=True, size=12)
                    current_row += 1
                    sheet.insert_rows(current_row)
                
                # Add content
                if content_type == 'bullet_list':
                    # Split into bullet points
                    lines = content.split('\n')
                    for line in lines:
                        if line.strip():
                            sheet.insert_rows(current_row)
                            content_cell = sheet.cell(current_row, 1)
                            content_cell.value = f"• {line.strip()}"
                            content_cell.font = Font(size=10)
                            current_row += 1
                else:
                    # Add as paragraphs
                    paragraphs = content.split('\n\n')
                    for para in paragraphs:
                        if para.strip():
                            sheet.insert_rows(current_row)
                            content_cell = sheet.cell(current_row, 1)
                            content_cell.value = para.strip()
                            content_cell.font = Font(size=10)
                            content_cell.alignment = Alignment(wrap_text=True)
                            current_row += 1
                
                # Add blank row after section
                sheet.insert_rows(current_row)
                current_row += 1
                
                sections_inserted += 1
                print(f"  ✓ Inserted: {title or 'Untitled section'}")
        
        print(f"✓ Inserted {sections_inserted} technical sections")
        print(f"=" * 60)

# Find the pricing table header row and map columns (AFTER insertions)
table_header_row = None
col_map = {}

for row_idx in range(1, min(50, sheet.max_row + 1)):
    row = sheet[row_idx]
    row_values = [str(cell.value).upper() if cell.value else '' for cell in row]
    row_text = ' '.join(row_values)
    
    if any(kw in row_text for kw in ['POSITION', 'DESCRIPTION', 'PRICE', 'QUANTITY', 'TOTAL']):
        table_header_row = row_idx
        print(f"✓ Found pricing table header at row {table_header_row}")
        
        # Map columns based on header text
        for col_idx, cell in enumerate(row, start=1):
            if not cell.value:
                continue
            header_text = str(cell.value).upper()
            
            if 'POSITION' in header_text or header_text.strip() == '#':
                col_map['position'] = col_idx
            elif 'DESCRIPTION' in header_text:
                col_map['description'] = col_idx
            elif 'UNIT' in header_text and 'PRICE' in header_text:
                col_map['unit_price'] = col_idx
            elif 'DISCOUNT' in header_text:
                col_map['discount'] = col_idx
            elif 'QUANTITY' in header_text or 'QTY' in header_text:
                col_map['quantity'] = col_idx
            elif 'TOTAL' in header_text and 'PRICE' in header_text:
                col_map['total_price'] = col_idx
        
        print(f"  Column mapping: {col_map}")
        break

if not table_header_row:
    print("✗ Could not find pricing table header")
    exit(1)

# Find where data rows start and end
data_start_row = table_header_row + 1
data_end_row = sheet.max_row

# Find actual end of data (last non-empty row in description column)
desc_col = col_map.get('description', 2)
for row_idx in range(sheet.max_row, data_start_row - 1, -1):
    cell_value = sheet.cell(row_idx, desc_col).value
    if cell_value and str(cell_value).strip():
        data_end_row = row_idx
        break

print(f"Data rows: {data_start_row} to {data_end_row}")

# Delete existing data rows (keep header and everything below the table)
rows_to_delete = data_end_row - data_start_row + 1
if rows_to_delete > 0:
    sheet.delete_rows(data_start_row, rows_to_delete)
    print(f"✓ Deleted {rows_to_delete} existing data rows")

# Group items by category
categorized_items = OrderedDict()
for item in items:
    category = item.get('category', 'Items')
    if category not in categorized_items:
        categorized_items[category] = []
    categorized_items[category].append(item)

print(f"✓ Grouped into {len(categorized_items)} categories")

# Insert items
current_row = data_start_row

for category, cat_items in categorized_items.items():
    # Add category header row
    sheet.insert_rows(current_row)
    
    # Put category name in description column, make it bold
    desc_col = col_map.get('description', 2)
    category_cell = sheet.cell(current_row, desc_col)
    category_cell.value = category
    category_cell.font = Font(bold=True, size=11)
    category_cell.alignment = Alignment(horizontal='left')
    
    current_row += 1
    
    # Add item rows
    for idx, item in enumerate(cat_items, start=1):
        sheet.insert_rows(current_row)
        
        # Position number
        if 'position' in col_map:
            sheet.cell(current_row, col_map['position']).value = str(idx)
        
        # Description
        if 'description' in col_map:
            # BUILD COMPLETE DESCRIPTION (SV12)
            desc_parts = []
            desc_parts.append(item.get('item_name', ''))
            
            if item.get('description'):
                desc_parts.append(item.get('description'))
            
            if item.get('specifications'):
                desc_parts.append(f"Specifications: {item.get('specifications')}")
            
            if item.get('has_image') and item.get('image_description'):
                desc_parts.append(f"[Image: {item.get('image_description')}]")
            
            if item.get('details'):
                desc_parts.append(item.get('details'))
            
            full_description = '\n\n'.join(desc_parts)
            
            desc_cell = sheet.cell(current_row, col_map['description'])
            desc_cell.value = full_description
            desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Unit Price
        if 'unit_price' in col_map:
            unit_price = item.get('unit_price', '')
            sheet.cell(current_row, col_map['unit_price']).value = unit_price
            sheet.cell(current_row, col_map['unit_price']).alignment = Alignment(horizontal='right')
        
        # Discount (leave empty)
        if 'discount' in col_map:
            sheet.cell(current_row, col_map['discount']).value = ''
        
        # Quantity
        if 'quantity' in col_map:
            qty = item.get('quantity', '1')
            sheet.cell(current_row, col_map['quantity']).value = qty
            sheet.cell(current_row, col_map['quantity']).alignment = Alignment(horizontal='center')
        
        # Total Price
        if 'total_price' in col_map:
            total = item.get('total_price', item.get('unit_price', ''))
            sheet.cell(current_row, col_map['total_price']).value = total
            sheet.cell(current_row, col_map['total_price']).alignment = Alignment(horizontal='right')
        
        current_row += 1

print(f"✓ Inserted {len(items)} items into Excel")

# Save output
try:
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    workbook.save(OUTPUT_PATH)
    file_size = os.path.getsize(OUTPUT_PATH)
    print(f"✓ Saved: {OUTPUT_PATH}")
    print(f"  File size: {file_size:,} bytes")
except Exception as e:
    print(f"✗ Error saving: {str(e)}")
    exit(1)

print("=" * 60)
print("GENERATION COMPLETE")
print("=" * 60)